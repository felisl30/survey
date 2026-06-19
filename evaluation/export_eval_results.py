#!/usr/bin/env python3
"""
export_eval_results.py

Une los resultados evaluados de S1/S2/S3 en un único XLSX con:
  - Una hoja por sistema con schema unificado
  - Hoja "resumen" con métricas agregadas por sistema y por tipo (retrieve/direct)
  - Hoja "metadata" con info de la corrida

Uso:
    python evaluation/export_eval_results.py --input-dir outputs/eval --model default
    python evaluation/export_eval_results.py --input-dir outputs/eval --model gpt-4o-mini
    python evaluation/export_eval_results.py --input-dir outputs/eval --model default --output mi_evaluacion.xlsx
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from project_paths import EVAL_OUTPUTS_DIR, EVAL_QUESTIONS_PATH

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# Columnas del schema unificado (orden de aparición en las hojas)
# ---------------------------------------------------------------------------

UNIFIED_COLUMNS = [
    "question_id",
    "question",
    "gold_answer",
    "expected_route",
    "source_dataset",
    "model_answer",
    "is_correct",
    "token_f1",
    "contains_gold",
    "latency_total_s",
    "latency_retrieval_s",
    "latency_generation_s",
    "latency_router_s",
    "n_docs_retrieved",
    "retrieved_chunk_ids",
    "retrieval_recall",
    "retrieval_precision",
    "predicted_route",
    "route_match_expected",
    "n_generation_steps",
    "n_retrieval_steps",
    "tokens_in",
    "tokens_out",
    "tokens_total",
    "confidence",
    "error",
]

# ---------------------------------------------------------------------------
# Mapeo de columnas por sistema hacia el schema unificado
# ---------------------------------------------------------------------------

def coerce_float(val: Any, default: float | None = None) -> float | None:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def coerce_int(val: Any, default: int | None = None) -> int | None:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def coerce_bool(val: Any) -> bool | None:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("true", "1", "yes")
    try:
        return bool(int(val))
    except (TypeError, ValueError):
        return None


def map_s1_row(row: pd.Series, questions: pd.DataFrame) -> dict:
    qid = str(row.get("id", ""))
    q_meta = questions[questions["id"] == qid].iloc[0] if qid in questions["id"].values else pd.Series()

    lat_total = coerce_float(row.get("row_wall_latency_seconds") or row.get("latency_seconds"))
    lat_ret = coerce_float(row.get("retrieval_latency_seconds"))
    lat_gen = coerce_float(row.get("generation_latency_seconds") or row.get("latency_seconds"))

    chunk_ids = str(row.get("retrieved_chunk_ids", "") or "")
    n_docs = len([c for c in chunk_ids.split("|") if c.strip()]) if chunk_ids else 0

    return {
        "question_id": qid,
        "question": str(q_meta.get("routing_question", q_meta.get("original_question", q_meta.get("question", "")))),
        "gold_answer": str(row.get("gold_answer", "")),
        "expected_route": str(q_meta.get("expected_route", "retrieve")),
        "source_dataset": str(q_meta.get("source_dataset", q_meta.get("dataset", ""))),
        "model_answer": str(row.get("parsed_answer", row.get("final_answer", ""))),
        "is_correct": coerce_bool(row.get("eval_answer_correct", row.get("eval_correct"))),
        "token_f1": coerce_float(row.get("eval_token_f1")),
        "contains_gold": coerce_bool(row.get("eval_contains_gold_answer")),
        "latency_total_s": lat_total,
        "latency_retrieval_s": lat_ret,
        "latency_generation_s": lat_gen,
        "latency_router_s": None,
        "n_docs_retrieved": n_docs,
        "retrieved_chunk_ids": chunk_ids,
        "retrieval_recall": coerce_float(row.get("eval_retrieval_recall")),
        "retrieval_precision": coerce_float(row.get("eval_retrieval_precision")),
        "predicted_route": None,
        "route_match_expected": None,
        "n_generation_steps": None,
        "n_retrieval_steps": None,
        "tokens_in": coerce_int(row.get("input_tokens")),
        "tokens_out": coerce_int(row.get("output_tokens")),
        "tokens_total": coerce_int(row.get("total_tokens")),
        "confidence": coerce_float(row.get("confidence")),
        "error": str(row.get("error", "") or ""),
    }


def map_s2_row(row: pd.Series, questions: pd.DataFrame) -> dict:
    qid = str(row.get("id", ""))
    q_meta = questions[questions["id"] == qid].iloc[0] if qid in questions["id"].values else pd.Series()

    lat_total = coerce_float(row.get("row_wall_latency_seconds") or row.get("latency_seconds"))
    lat_ret = coerce_float(row.get("retrieval_latency_seconds"))
    lat_gen = coerce_float(row.get("generation_latency_seconds") or row.get("latency_seconds"))
    lat_router = coerce_float(row.get("router_latency_seconds"))

    chunk_ids = str(row.get("retrieved_chunk_ids", "") or "")
    n_docs = len([c for c in chunk_ids.split("|") if c.strip()]) if chunk_ids else 0

    predicted = str(row.get("predicted_route", "") or "")
    expected = str(q_meta.get("expected_route", "") or row.get("expected_route", "") or "")
    route_match = (predicted == expected) if predicted and expected else None

    tokens_in = coerce_int(row.get("input_tokens") or
                            (coerce_int(row.get("router_input_tokens", 0) or 0) +
                             coerce_int(row.get("generation_input_tokens", 0) or 0)))
    tokens_out = coerce_int(row.get("output_tokens") or
                             (coerce_int(row.get("router_output_tokens", 0) or 0) +
                              coerce_int(row.get("generation_output_tokens", 0) or 0)))
    tokens_total = coerce_int(row.get("total_tokens"))
    if tokens_total is None and tokens_in is not None and tokens_out is not None:
        tokens_total = tokens_in + tokens_out

    return {
        "question_id": qid,
        "question": str(q_meta.get("routing_question", q_meta.get("original_question", q_meta.get("question", "")))),
        "gold_answer": str(row.get("gold_answer", "")),
        "expected_route": expected,
        "source_dataset": str(q_meta.get("source_dataset", q_meta.get("dataset", ""))),
        "model_answer": str(row.get("parsed_answer", row.get("final_answer", ""))),
        "is_correct": coerce_bool(row.get("eval_answer_correct", row.get("eval_correct"))),
        "token_f1": coerce_float(row.get("eval_token_f1")),
        "contains_gold": coerce_bool(row.get("eval_contains_gold_answer")),
        "latency_total_s": lat_total,
        "latency_retrieval_s": lat_ret,
        "latency_generation_s": lat_gen,
        "latency_router_s": lat_router,
        "n_docs_retrieved": n_docs,
        "retrieved_chunk_ids": chunk_ids,
        "retrieval_recall": coerce_float(row.get("eval_retrieval_recall")),
        "retrieval_precision": coerce_float(row.get("eval_retrieval_precision")),
        "predicted_route": predicted,
        "route_match_expected": route_match,
        "n_generation_steps": None,
        "n_retrieval_steps": None,
        "tokens_in": tokens_in,
        "tokens_out": tokens_out,
        "tokens_total": tokens_total,
        "confidence": coerce_float(row.get("confidence")),
        "error": str(row.get("error", "") or ""),
    }


def map_s3_row(row: pd.Series, questions: pd.DataFrame) -> dict:
    qid = str(row.get("id", ""))
    q_meta = questions[questions["id"] == qid].iloc[0] if qid in questions["id"].values else pd.Series()

    lat_total = coerce_float(row.get("row_wall_latency_seconds") or row.get("latency_seconds"))
    lat_ret = coerce_float(row.get("retrieval_latency_seconds"))
    lat_gen = coerce_float(row.get("generation_latency_seconds") or row.get("latency_seconds"))

    chunk_ids_raw = str(row.get("retrieved_chunk_ids_json", row.get("retrieved_chunk_ids", "")) or "")
    if chunk_ids_raw.startswith("["):
        try:
            ids_list = json.loads(chunk_ids_raw)
            if isinstance(ids_list, list):
                chunk_ids = "|".join(str(x) for x in ids_list)
                n_docs = len(ids_list)
            else:
                chunk_ids = chunk_ids_raw
                n_docs = coerce_int(row.get("num_chunks_retrieved_total", row.get("eval_chunks_retrieved"))) or 0
        except json.JSONDecodeError:
            chunk_ids = chunk_ids_raw
            n_docs = coerce_int(row.get("num_chunks_retrieved_total", row.get("eval_chunks_retrieved"))) or 0
    else:
        chunk_ids = chunk_ids_raw
        n_docs = coerce_int(row.get("num_chunks_retrieved_total", row.get("eval_chunks_retrieved"))) or 0

    return {
        "question_id": qid,
        "question": str(q_meta.get("routing_question", q_meta.get("original_question", q_meta.get("question", "")))),
        "gold_answer": str(row.get("gold_answer", "")),
        "expected_route": str(q_meta.get("expected_route", "retrieve")),
        "source_dataset": str(q_meta.get("source_dataset", q_meta.get("dataset", ""))),
        "model_answer": str(row.get("parsed_answer", row.get("final_answer", ""))),
        "is_correct": coerce_bool(row.get("eval_answer_correct", row.get("eval_correct"))),
        "token_f1": coerce_float(row.get("eval_token_f1")),
        "contains_gold": coerce_bool(row.get("eval_contains_gold_answer")),
        "latency_total_s": lat_total,
        "latency_retrieval_s": lat_ret,
        "latency_generation_s": lat_gen,
        "latency_router_s": None,
        "n_docs_retrieved": n_docs,
        "retrieved_chunk_ids": chunk_ids,
        "retrieval_recall": coerce_float(row.get("eval_retrieval_recall")),
        "retrieval_precision": coerce_float(row.get("eval_retrieval_precision")),
        "predicted_route": None,
        "route_match_expected": None,
        "n_generation_steps": coerce_int(row.get("num_generation_steps")),
        "n_retrieval_steps": coerce_int(row.get("num_retrieval_steps")),
        "tokens_in": coerce_int(row.get("input_tokens")),
        "tokens_out": coerce_int(row.get("output_tokens")),
        "tokens_total": coerce_int(row.get("total_tokens")),
        "confidence": coerce_float(row.get("final_confidence", row.get("confidence"))),
        "error": str(row.get("error", "") or ""),
    }


SYSTEM_MAPPERS = {
    "s1": map_s1_row,
    "s2": map_s2_row,
    "s3": map_s3_row,
}


# ---------------------------------------------------------------------------
# Build per-system DataFrame
# ---------------------------------------------------------------------------

def load_system_df(results_path: Path, system: str, questions: pd.DataFrame) -> pd.DataFrame:
    df = pd.read_csv(results_path)
    mapper = SYSTEM_MAPPERS[system]
    rows = [mapper(row, questions) for _, row in df.iterrows()]
    return pd.DataFrame(rows, columns=UNIFIED_COLUMNS)


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _safe_mean(series: pd.Series) -> float | None:
    s = pd.to_numeric(series, errors="coerce").dropna()
    return float(s.mean()) if len(s) > 0 else None


def _safe_rate(series: pd.Series) -> float | None:
    s = series.dropna()
    if len(s) == 0:
        return None
    return float((s == True).sum() / len(s))  # noqa: E712


def build_summary_df(system_dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for system, df in system_dfs.items():
        for route_group, label in [
            (None, "ALL"),
            ("retrieve", "retrieve"),
            ("direct", "direct"),
        ]:
            sub = df if route_group is None else df[df["expected_route"] == route_group]
            if len(sub) == 0:
                continue

            row: dict[str, Any] = {
                "sistema": system.upper(),
                "grupo": label,
                "n": len(sub),
                "accuracy": _safe_rate(sub["is_correct"]),
                "avg_token_f1": _safe_mean(sub["token_f1"]),
                "contains_gold_rate": _safe_rate(sub["contains_gold"]),
                "avg_latency_total_s": _safe_mean(sub["latency_total_s"]),
                "avg_latency_retrieval_s": _safe_mean(sub["latency_retrieval_s"]),
                "avg_latency_generation_s": _safe_mean(sub["latency_generation_s"]),
                "avg_n_docs": _safe_mean(pd.to_numeric(sub["n_docs_retrieved"], errors="coerce")),
                "avg_tokens_total": _safe_mean(sub["tokens_total"]),
                "avg_retrieval_recall": _safe_mean(sub["retrieval_recall"]),
                "avg_retrieval_precision": _safe_mean(sub["retrieval_precision"]),
                "error_rate": _safe_rate(sub["error"].apply(lambda x: bool(x and str(x).strip() not in ("", "nan")))),
            }

            if system == "s2" and "route_match_expected" in sub.columns:
                row["routing_accuracy"] = _safe_rate(sub["route_match_expected"])
            else:
                row["routing_accuracy"] = None

            if system == "s3":
                row["avg_generation_steps"] = _safe_mean(sub["n_generation_steps"])
                row["avg_retrieval_steps"] = _safe_mean(sub["n_retrieval_steps"])
            else:
                row["avg_generation_steps"] = None
                row["avg_retrieval_steps"] = None

            rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def write_xlsx(
    system_dfs: dict[str, pd.DataFrame],
    summary_df: pd.DataFrame,
    output_path: Path,
    metadata: dict,
) -> None:
    if not OPENPYXL_OK:
        raise ImportError("Instalá openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="D9E1F2")

    def style_header(ws, n_cols: int) -> None:
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def write_df(ws, df: pd.DataFrame) -> None:
        for ci, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=ci, value=col)
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
        style_header(ws, len(df.columns))
        auto_width(ws)

    for system, df in system_dfs.items():
        ws = wb.create_sheet(title=system.upper())
        write_df(ws, df)

    ws_sum = wb.create_sheet(title="resumen")
    write_df(ws_sum, summary_df)

    ws_meta = wb.create_sheet(title="metadata")
    ws_meta.cell(row=1, column=1, value="clave")
    ws_meta.cell(row=1, column=2, value="valor")
    style_header(ws_meta, 2)
    for i, (k, v) in enumerate(metadata.items(), start=2):
        ws_meta.cell(row=i, column=1, value=str(k))
        ws_meta.cell(row=i, column=2, value=str(v))
    auto_width(ws_meta)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"XLSX guardado: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exporta resultados de evaluación S1/S2/S3 a un XLSX unificado."
    )
    parser.add_argument("--input-dir", type=Path, default=EVAL_OUTPUTS_DIR,
                        help="Directorio con los CSVs evaluados (salida de run_evaluation_pipeline.py).")
    parser.add_argument("--model", type=str, default="default",
                        help="Tag del modelo (para encontrar los archivos sNN_<model>_results.csv).")
    parser.add_argument("--systems", type=str, default="s1,s2,s3",
                        help="Sistemas a incluir. Default: s1,s2,s3.")
    parser.add_argument("--questions-path", type=Path, default=EVAL_QUESTIONS_PATH,
                        help="CSV de preguntas (para lookup de metadata como expected_route).")
    parser.add_argument("--output", type=Path, default=None,
                        help="Ruta del XLSX de salida. Default: <input-dir>/evaluation_<model>.xlsx")
    args = parser.parse_args()

    model_tag = args.model.replace("/", "_").replace("-", "_").replace(".", "_")
    systems = [s.strip().lower() for s in args.systems.split(",")]

    output_path = args.output or (args.input_dir / f"evaluation_{model_tag}.xlsx")

    questions = pd.read_csv(args.questions_path) if args.questions_path.exists() else pd.DataFrame()

    system_dfs: dict[str, pd.DataFrame] = {}
    for system in systems:
        results_path = args.input_dir / f"{system}_{model_tag}_results.csv"
        if not results_path.exists():
            print(f"[SKIP] No existe: {results_path}")
            continue
        print(f"Cargando {system.upper()}: {results_path}")
        system_dfs[system] = load_system_df(results_path, system, questions)

    if not system_dfs:
        print("No se encontraron archivos de resultados. Verificá --input-dir y --model.")
        sys.exit(1)

    summary_df = build_summary_df(system_dfs)

    try:
        commit = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"], stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        commit = "unknown"

    metadata = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "model": args.model,
        "systems": ", ".join(system_dfs),
        "total_questions": sum(len(df) for df in system_dfs.values()) // max(len(system_dfs), 1),
        "input_dir": str(args.input_dir),
        "questions_path": str(args.questions_path),
        "git_commit": commit,
    }
    for system, df in system_dfs.items():
        metadata[f"n_{system}"] = len(df)

    write_xlsx(system_dfs, summary_df, output_path, metadata)

    print("\n=== Resumen por sistema ===")
    print(summary_df[["sistema", "grupo", "n", "accuracy", "avg_token_f1",
                        "avg_latency_total_s", "avg_n_docs"]].to_string(index=False))


if __name__ == "__main__":
    main()
